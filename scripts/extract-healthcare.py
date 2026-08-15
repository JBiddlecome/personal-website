"""Extract healthcare rows from Denver_Healthcare_Market_Research.xlsx into scripts/healthcare-rows.json.
Skips 'Outside metro' LTC facilities and metro-wide aggregate rows in Other Facilities.
Usage: python scripts/extract-healthcare.py [path-to-xlsx]
"""
import json, os, sys, shutil, tempfile
import openpyxl

src = sys.argv[1] if len(sys.argv) > 1 else os.path.join(os.path.dirname(__file__), '..', 'Denver_Healthcare_Market_Research.xlsx')
tmp = os.path.join(tempfile.gettempdir(), 'hc_extract.xlsx')
shutil.copy(src, tmp)  # OneDrive sometimes locks the original
wb = openpyxl.load_workbook(tmp, data_only=True)

rows = []

# ── Hospitals ──
ws = wb['Hospitals']
for r in ws.iter_rows(min_row=4, values_only=True):
    if not isinstance(r[0], int):
        continue
    n, name, city, parent, beds, typ, region, notes = r[:8]
    if 'hospital-within-hospital' in (notes or '').lower():
        continue  # Rocky Mountain Children's is inside P/SL
    rows.append(dict(type='hospital', name=name, city=city, parent=parent, beds=beds, care=typ, region=region, notes=notes or ''))

# ── LTC Facilities ──
ws = wb['LTC Facilities']
for r in ws.iter_rows(min_row=4, values_only=True):
    if not isinstance(r[0], int):
        continue
    n, name, city, op, beds, care, region, notes = r[:8]
    if region == 'Outside metro':
        continue
    if 'IL-only communities' in name:
        continue  # aggregate row
    care_l = (care or '').lower()
    t = 'snf' if care_l.startswith('skilled') or care_l.startswith('snf') else 'senior'
    rows.append(dict(type=t, name=name, city=city, parent=op, beds=beds, care=care, region=region, notes=notes or ''))

# ── Other Facilities (single-site only) ──
ws = wb['Other Facilities']
TYPE_BY_CAT = {
    'Behavioral health / psychiatric hospitals': 'behavioral',
    'Inpatient rehabilitation hospitals': 'rehab',
    'Long-term acute care hospitals (LTACH)': 'rehab',
}
SINGLE_SITE_OTHER = {
    # name -> (type, address hint)
    'DaVita Inc.': ('other', '2000 16th St, Denver, CO 80202'),
    'The Denver Hospice': ('other', '8289 E Lowry Blvd, Denver, CO 80230'),
    'WellPower (formerly Mental Health Center of Denver)': ('behavioral', '4141 E Dickenson Pl, Denver, CO 80222'),
    'Jefferson Center for Mental Health': ('behavioral', '4851 Independence St, Wheat Ridge, CO 80033'),
    'Mental Health Partners': ('behavioral', '1333 Iris Ave, Boulder, CO 80304'),
    'AllHealth Network': ('behavioral', '6509 S Santa Fe Dr, Littleton, CO 80120'),
    'Community Reach Center': ('behavioral', '8931 Huron St, Thornton, CO 80260'),
    'Fitzsimons Innovation Community': ('other', '12635 E Montview Blvd, Aurora, CO 80045'),
    'Vitalant (formerly Bonfils Blood Center)': ('other', '717 Yosemite St, Denver, CO 80230'),
    'Colorado Coalition for the Homeless': ('other', '2130 Stout St, Denver, CO 80205'),
    'Clinica Family Health & Wellness': ('other', '1735 S Public Rd, Lafayette, CO 80026'),
    'Salud Family Health Centers': ('other', '203 S Rollie Ave, Fort Lupton, CO 80621'),
    'Boulder Community Health - Community Medical Center ED': ('other', '1000 W South Boulder Rd, Lafayette, CO 80026'),
    'HCA HealthONE Ambulatory Surgical Centers': ('other', '4900 S Monaco St, Denver, CO 80237'),
    'TRU Community Care': ('other', '2594 Trailridge Dr E, Lafayette, CO 80026'),
    'HCA HealthONE Hospice & Family Care': ('other', '6600 S Syracuse Way, Greenwood Village, CO 80111'),
}
for r in ws.iter_rows(min_row=4, values_only=True):
    if not isinstance(r[0], int):
        continue
    n, name, city, parent, size, unit, cat, notes = r[:8]
    if cat in TYPE_BY_CAT:
        if name in ('Craig Hospital', 'HCA HealthONE Spalding Rehabilitation', 'Colorado Mental Health Hospital in Fort Logan'):
            # Craig/Spalding already on Hospitals tab; Fort Logan already pinned under Government > State
            continue
        rows.append(dict(type=TYPE_BY_CAT[cat], name=name, city=city, parent=parent, beds=size, care=cat, region='Metro', notes=notes or ''))
    elif name in SINGLE_SITE_OTHER:
        t, addr = SINGLE_SITE_OTHER[name]
        stat = f"{size} {unit}" if size else (unit or '')
        rows.append(dict(type=t, name=name, city=city, parent=parent, beds=None, stat=stat, care=cat, region='Metro', notes=notes or '', address=addr))

out = os.path.join(os.path.dirname(__file__), 'healthcare-rows.json')
json.dump(rows, open(out, 'w', encoding='utf8'), indent=1, ensure_ascii=False)
from collections import Counter
print(len(rows), Counter(r['type'] for r in rows))
